"""
Testes para o módulo de geração de Excel (excel.py).

Cobertura:
- create_excel() com diferentes inputs (vazio, único, múltiplos)
- Validação de formatação (cores, fontes, bordas)
- Parsing de datas em múltiplos formatos
- Metadados e fórmulas
- Edge cases (valores None, strings vazias)
"""

from contextlib import contextmanager
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from excel import create_excel, parse_datetime


@contextmanager
def open_workbook(buffer: BytesIO):
    """Context manager para garantir fechamento do workbook."""
    wb = load_workbook(buffer)
    try:
        yield wb
    finally:
        wb.close()


class TestParseDateTime:
    """Testes para parse_datetime()."""

    def test_parse_iso_with_z_timezone(self):
        """Deve parsear ISO 8601 com Z (UTC)."""
        result = parse_datetime("2024-01-25T10:30:00Z")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_iso_with_offset_timezone(self):
        """Deve parsear ISO 8601 com offset (+00:00)."""
        result = parse_datetime("2024-01-25T10:30:00+00:00")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_iso_without_timezone(self):
        """Deve parsear ISO 8601 sem timezone."""
        result = parse_datetime("2024-01-25T10:30:00")
        assert result == datetime(2024, 1, 25, 10, 30, 0)

    def test_parse_date_only(self):
        """Deve parsear data simples (YYYY-MM-DD)."""
        result = parse_datetime("2024-01-25")
        assert result == datetime(2024, 1, 25, 0, 0, 0)

    def test_parse_none(self):
        """Deve retornar None para valor None."""
        assert parse_datetime(None) is None

    def test_parse_empty_string(self):
        """Deve retornar None para string vazia."""
        assert parse_datetime("") is None

    def test_parse_invalid_format(self):
        """Deve retornar None para formato inválido."""
        assert parse_datetime("25/01/2024") is None
        assert parse_datetime("not a date") is None


class TestCreateExcel:
    """Testes para create_excel()."""

    def test_create_excel_returns_bytesio(self):
        """Deve retornar BytesIO buffer."""
        result = create_excel([])
        assert isinstance(result, BytesIO)

    def test_create_excel_with_empty_list(self):
        """Deve gerar Excel válido mesmo com lista vazia."""
        buffer = create_excel([])

        with open_workbook(buffer) as wb:
            # Verificar sheet principal existe
            assert "Licitações Uniformes" in wb.sheetnames
            ws = wb["Licitações Uniformes"]

            # Verificar headers
            assert ws["A1"].value == "Tipo"
            assert ws["B1"].value == "Objeto"
            assert ws["C1"].value == "Órgão"
            assert ws["J1"].value == "Link"

            # Verificar formatação do header
            assert ws["A1"].font.bold is True
            assert (
                ws["A1"].font.color.rgb == "00FFFFFF"
            )  # Branco (openpyxl ARGB format)
            assert (
                ws["A1"].fill.start_color.rgb == "002E7D32"
            )  # Verde (openpyxl ARGB format)

            # Verificar que não há linha de totais (lista vazia)
            assert ws["F3"].value is None

    def test_create_excel_with_single_item(self):
        """Deve gerar Excel com um item e linha de totais."""
        licitacao = {
            "codigoCompra": "12345",
            "objetoCompra": "Aquisição de uniformes escolares",
            "nomeOrgao": "Prefeitura Municipal",
            "uf": "SP",
            "municipio": "São Paulo",
            "valorTotalEstimado": 150000.50,
            "modalidadeNome": "Pregão Eletrônico",
            "dataPublicacaoPncp": "2024-01-20T08:00:00Z",
            "dataAberturaProposta": "2024-02-01T10:00:00Z",
            "situacaoCompraNome": "Em Disputa",
            "linkPncp": "https://pncp.gov.br/app/editais/12345",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar dados na linha 2
            assert ws["A2"].value == "Licitacao"  # Tipo column
            assert ws["B2"].value == "Aquisição de uniformes escolares"
            assert ws["C2"].value == "Prefeitura Municipal"
            assert ws["D2"].value == "SP"
            assert ws["E2"].value == "São Paulo"
            assert ws["F2"].value == 150000.50
            assert ws["G2"].value == "Pregão Eletrônico"

            # Verificar formatação de moeda
            assert "R$" in ws["F2"].number_format

            # Verificar hyperlink (coluna J)
            assert ws["J2"].value == "Abrir"
            assert ws["J2"].hyperlink.target == "https://pncp.gov.br/app/editais/12345"
            assert ws["J2"].font.color.rgb == "000563C1"  # Azul (openpyxl ARGB format)
            assert ws["J2"].font.underline == "single"

            # Verificar linha de totais (row 3)
            assert ws["E3"].value == "TOTAL:"
            assert ws["E3"].font.bold is True
            assert "=SUM(F2:F2)" in ws["F3"].value
            assert ws["F3"].font.bold is True

            # Verificar bordas
            assert ws["A2"].border.left.style == "thin"

    def test_create_excel_with_multiple_items(self):
        """Deve gerar Excel com múltiplos itens."""
        licitacoes = [
            {
                "codigoCompra": "123",
                "objetoCompra": "Item 1",
                "valorTotalEstimado": 100000,
            },
            {
                "codigoCompra": "456",
                "objetoCompra": "Item 2",
                "valorTotalEstimado": 200000,
            },
            {
                "codigoCompra": "789",
                "objetoCompra": "Item 3",
                "valorTotalEstimado": 300000,
            },
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Verificar 3 linhas de dados
        assert ws["B2"].value == "Item 1"
        assert ws["B3"].value == "Item 2"
        assert ws["B4"].value == "Item 3"

        # Verificar linha de totais (row 5)
        assert "=SUM(F2:F4)" in ws["F5"].value

    def test_create_excel_with_none_values(self):
        """Deve lidar com valores None nos campos."""
        licitacao = {
            "codigoCompra": "999",
            "objetoCompra": None,
            "valorTotalEstimado": None,
            "dataPublicacaoPncp": None,
            "dataAberturaProposta": None,
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Verificar que não crashou
        assert ws["A2"].value == "Licitacao"  # Tipo column (always present)
        assert ws["B2"].value is None  # Objeto
        assert ws["F2"].value is None  # Valor
        assert ws["H2"].value is None  # Publicacao
        assert ws["I2"].value is None  # Início

    def test_create_excel_with_missing_link(self):
        """Deve gerar link padrão usando CNPJ/ano/sequencial se linkPncp não existir."""
        licitacao = {
            "codigoCompra": "83845701000159-1-000016/2026",
            "cnpj": "83845701000159",
            "anoCompra": 2026,
            "sequencialCompra": 16,
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Verificar link gerado com formato correto: /editais/{cnpj}/{ano}/{seq}
        assert ws["J2"].hyperlink.target == "https://pncp.gov.br/app/editais/83845701000159/2026/16"

    def test_create_excel_fallback_link_without_cnpj(self):
        """Deve gerar link de busca se CNPJ/ano/sequencial não existirem."""
        licitacao = {"codigoCompra": "ABC123"}

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

        # Fallback: search URL when structured fields are missing
        assert ws["J2"].hyperlink.target == "https://pncp.gov.br/app/editais?q=ABC123"

    def test_create_excel_frozen_panes(self):
        """Deve congelar a primeira linha (headers)."""
        buffer = create_excel([{"codigoCompra": "123"}])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar frozen panes
            assert ws.freeze_panes == "A2"

    def test_create_excel_metadata_sheet(self):
        """Deve criar aba de Metadata com estatísticas."""
        licitacoes = [
            {"valorTotalEstimado": 100000},
            {"valorTotalEstimado": 200000},
            {"valorTotalEstimado": 300000},
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            # Verificar que existe aba Metadata
            assert "Metadata" in wb.sheetnames
            ws_meta = wb["Metadata"]

            # Verificar conteúdo
            assert ws_meta["A1"].value == "Gerado em:"
            assert ws_meta["B1"].value is not None  # Timestamp

            assert ws_meta["A2"].value == "Total de licitações:"
            assert ws_meta["B2"].value == 3

            assert ws_meta["A3"].value == "Valor total estimado:"
            assert ws_meta["B3"].value == 600000
            assert "R$" in ws_meta["B3"].number_format

    def test_create_excel_metadata_with_none_values(self):
        """Metadata deve somar corretamente mesmo com None."""
        licitacoes = [
            {"valorTotalEstimado": 100000},
            {"valorTotalEstimado": None},  # Deve ignorar
            {"valorTotalEstimado": 50000},
        ]

        buffer = create_excel(licitacoes)

        with open_workbook(buffer) as wb:
            ws_meta = wb["Metadata"]

            assert ws_meta["B2"].value == 3  # 3 licitações
            assert ws_meta["B3"].value == 150000  # Soma ignora None

    def test_create_excel_date_formatting(self):
        """Deve aplicar formatação correta nas colunas de data."""
        licitacao = {
            "dataPublicacaoPncp": "2024-01-25T10:00:00Z",
            "dataAberturaProposta": "2024-02-01T14:30:00Z",
        }

        buffer = create_excel([licitacao])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar formatação de data (columns shifted +1 for Tipo)
            assert "DD/MM/YYYY" in ws["H2"].number_format  # Publicação
            assert "DD/MM/YYYY HH:MM" in ws["I2"].number_format  # Início

    def test_create_excel_column_widths(self):
        """Deve definir larguras corretas das colunas."""
        buffer = create_excel([])

        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            # Verificar algumas larguras (PRD Section 5.1)
            assert ws.column_dimensions["A"].width == 12  # Tipo
            assert ws.column_dimensions["B"].width == 60  # Objeto
            assert ws.column_dimensions["C"].width == 40  # Órgão
            assert ws.column_dimensions["D"].width == 6  # UF

    def test_create_excel_invalid_input(self):
        """Deve lançar ValueError se input não for lista."""
        with pytest.raises(ValueError, match="licitacoes deve ser uma lista"):
            create_excel("not a list")

        with pytest.raises(ValueError):
            create_excel({"not": "a list"})

    def test_create_excel_can_be_reopened(self):
        """Buffer gerado deve ser um Excel válido que pode ser reaberto."""
        licitacao = {"codigoCompra": "TEST123", "objetoCompra": "Test"}
        buffer = create_excel([licitacao])

        # Salvar e reabrir para validar integridade
        with open_workbook(buffer) as wb:
            ws = wb["Licitações Uniformes"]

            assert ws["A2"].value == "Licitacao"  # Tipo column
            assert ws["B2"].value == "Test"  # Objeto column

            # Verificar que pode ser salvo novamente
            new_buffer = BytesIO()
            wb.save(new_buffer)
            new_buffer.seek(0)

        with open_workbook(new_buffer) as wb2:
            assert "Licitações Uniformes" in wb2.sheetnames
