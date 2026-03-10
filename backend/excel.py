"""
Gerador de planilhas Excel formatadas para licitações de uniformes.

Este módulo é responsável por criar arquivos Excel profissionalmente formatados
com as licitações filtradas, incluindo:
- Header verde (#2E7D32) com texto branco
- 11 colunas com larguras otimizadas
- Formatação de moeda (R$), datas e hyperlinks
- Linha de totais com fórmula SUM
- Aba de metadados com estatísticas da busca

Exemplo de uso:
    >>> from excel import create_excel
    >>> licitacoes = [{"codigoCompra": "123", "objetoCompra": "Uniformes", ...}]
    >>> buffer = create_excel(licitacoes)
    >>> with open("output.xlsx", "wb") as f:
    ...     f.write(buffer.getvalue())
"""

import re
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Regex to strip illegal XML 1.0 control characters that openpyxl rejects.
# Keeps tab (\x09), newline (\x0a), and carriage return (\x0d).
_ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]")


def _sanitize(value: str | None) -> str:
    """Remove illegal XML control characters from a string."""
    if not value:
        return ""
    return _ILLEGAL_XML_CHARS_RE.sub("", value)


def create_excel(licitacoes: list[dict]) -> BytesIO:
    """
    Gera planilha Excel formatada com licitações filtradas.

    Args:
        licitacoes: Lista de dicionários com dados das licitações do PNCP

    Returns:
        BytesIO: Buffer com arquivo Excel pronto para download/salvamento

    Raises:
        ValueError: Se licitacoes não for uma lista
    """
    if not isinstance(licitacoes, list):
        raise ValueError("licitacoes deve ser uma lista")

    wb = Workbook()
    ws = wb.active
    ws.title = "Licitações Uniformes"

    # === ESTILOS ===
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2E7D32", end_color="2E7D32", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    currency_format = '[$R$-416] #.##0,00'
    date_format = "DD/MM/YYYY"
    datetime_format = "DD/MM/YYYY HH:MM"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # === HEADERS ===
    headers = [
        ("Tipo", 12),
        ("Objeto", 60),
        ("Órgão", 40),
        ("UF", 6),
        ("Município", 20),
        ("Valor Estimado", 18),
        ("Modalidade", 20),
        ("Publicação", 12),
        ("Início", 16),
        ("Link", 15),
    ]

    for col, (header_name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Congelar header
    ws.freeze_panes = "A2"

    # === DADOS ===
    for row_idx, lic in enumerate(licitacoes, start=2):
        # A: Tipo
        tipo = lic.get("tipo", "licitacao")
        tipo_label = "Ata RP" if tipo == "ata_registro_preco" else "Licitacao"
        ws.cell(row=row_idx, column=1, value=tipo_label)

        # B: Objeto
        ws.cell(row=row_idx, column=2, value=_sanitize(lic.get("objetoCompra", "")))

        # C: Órgão
        ws.cell(row=row_idx, column=3, value=_sanitize(lic.get("nomeOrgao", "")))

        # D: UF
        ws.cell(row=row_idx, column=4, value=lic.get("uf", ""))

        # E: Município
        ws.cell(row=row_idx, column=5, value=_sanitize(lic.get("municipio", "")))

        # F: Valor (formatado como moeda)
        valor_cell = ws.cell(row=row_idx, column=6, value=lic.get("valorTotalEstimado"))
        valor_cell.number_format = currency_format

        # G: Modalidade
        ws.cell(row=row_idx, column=7, value=_sanitize(lic.get("modalidadeNome", "")))

        # H: Data Publicação
        data_pub = parse_datetime(lic.get("dataPublicacaoPncp"))
        pub_cell = ws.cell(row=row_idx, column=8, value=data_pub)
        if data_pub:
            pub_cell.number_format = date_format

        # I: Data Início (Abertura)
        data_abertura = parse_datetime(lic.get("dataAberturaProposta"))
        abertura_cell = ws.cell(row=row_idx, column=9, value=data_abertura)
        if data_abertura:
            abertura_cell.number_format = datetime_format

        # J: Link (hyperlink)
        cnpj = lic.get("cnpj", "")
        ano = lic.get("anoCompra", "")
        seq = lic.get("sequencialCompra", "")
        link = lic.get("linkPncp") or (
            f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{seq}"
            if cnpj and ano and seq
            else f"https://pncp.gov.br/app/editais?q={lic.get('codigoCompra', '')}"
        )
        link_cell = ws.cell(row=row_idx, column=10, value="Abrir")
        link_cell.hyperlink = link
        link_cell.font = Font(color="0563C1", underline="single")

        # Aplicar bordas e alinhamento em todas as células da linha
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = cell_alignment

    # === LINHA DE TOTAIS ===
    if licitacoes:  # Só adiciona linha de totais se houver dados
        total_row = len(licitacoes) + 2
        ws.cell(row=total_row, column=5, value="TOTAL:").font = Font(bold=True)

        total_cell = ws.cell(
            row=total_row, column=6, value=f"=SUM(F2:F{total_row - 1})"
        )
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)

    # === METADATA (aba separada) ===
    ws_meta = wb.create_sheet("Metadata")
    ws_meta["A1"] = "Gerado em:"
    ws_meta["B1"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws_meta["A2"] = "Total de licitações:"
    ws_meta["B2"] = len(licitacoes)
    ws_meta["A3"] = "Valor total estimado:"
    ws_meta["B3"] = sum(lic.get("valorTotalEstimado", 0) or 0 for lic in licitacoes)
    ws_meta["B3"].number_format = currency_format

    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def parse_datetime(value: str | None) -> datetime | None:
    """
    Parse datetime string do formato PNCP para objeto datetime.

    Tenta múltiplos formatos comuns:
    - ISO 8601 com timezone (2024-01-25T10:30:00Z)
    - ISO 8601 sem timezone (2024-01-25T10:30:00)
    - Data simples (2024-01-25)

    IMPORTANTE: Excel não suporta timezone-aware datetimes, então sempre
    retornamos naive datetime (tzinfo=None) mesmo para strings com timezone.

    Args:
        value: String de data/datetime do PNCP

    Returns:
        datetime object ou None se parsing falhar

    Examples:
        >>> parse_datetime("2024-01-25T10:30:00Z")
        datetime(2024, 1, 25, 10, 30, 0)
        >>> parse_datetime("2024-01-25")
        datetime(2024, 1, 25, 0, 0, 0)
        >>> parse_datetime(None)
        None
    """
    if not value:
        return None

    try:
        # Formato ISO com timezone (Z = UTC)
        # Excel não suporta timezone, então removemos tzinfo
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        pass

    try:
        # Formato sem timezone
        return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S")
    except (ValueError, AttributeError):
        pass

    try:
        # Apenas data
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, AttributeError):
        return None
